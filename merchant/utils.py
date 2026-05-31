import io
import re
import pandas as pd
import plotly.graph_objects as go
import plotly.offline as pyo

# =============================================================================
# CONSTANTS
# =============================================================================

CIF_BIZUNIT_OVERRIDES = {
    '004860': 'RETAIL',
}

CIF_GROUPS = [
    {
        'cifs': ['013714', '014345'],
        'primary_cif': '013714',
        'merchant': 'SLICE GROUP',
    },
]

CIF_NAME_OVERRIDES = {
    '013714': 'CHICKEN SLICE',
    '014345': 'CHICKEN SLICE',
    '106139': 'BLUE LAGOON',
    '106225': 'POTE',
    '106618': 'VILLAGE TECH',
    '279551': 'TYNWALD MARKET',
    '014911': 'CLUB LABAMBA',
    '004860': 'BANC ASSURANCE',
    '012429': 'SIMBISA',
}

MONTH_ABBR = {
    'JAN': 1, 'FEB': 2, 'MAR': 3, 'APR': 4, 'MAY': 5, 'JUN': 6,
    'JUL': 7, 'AUG': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DEC': 12,
}

BA_TIDS = [
    'NBSX1019', 'NBSX1012', 'NBSX1101', 'NBSX1102', 'NBSX1103',
    'NBSX1104', 'NBSX1105', 'NBSX1119', 'NBSX1538', 'NBSX1574',
]
ZT_TIDS = [
    'NBSX0435', 'NBSX0452', 'NBSX0434', 'NBSX0427', 'NBSX0429', 'NBSX0447',
    'NBSX0453', 'NBSX0437', 'NBSX0053', 'NBS30025', 'NBSX0433',
]

CIF_SHEET_NAMES = {
    '012765': 'CHINGWANGA',
    '013027': 'CITY OF CHINHOYI',
    '012647': 'CITY OF MASVINGO',
    '014557': 'CLARION',
    '013009': 'CREDSURE',
    '107251': 'CLEAR CHOICE',
    '014800': 'ENPASENT MULTI-AGENT',
    '010733': 'GREENWOOD PHARMACIES',
    '013364': 'HOUSE & HOMES',
    '010256': 'NETONE',
    '012852': 'OK ZIMBABWE',
    '013893': 'PG MERCH-TIMB-ZIMTILE',
    '012429': 'SIMBISA',
    '012814': 'SILO',
    '011251': 'TENDY THREE INVESTMENTS',
    '012913': 'ZIMNAT LION INSURANCE',
    '012608': 'SEEDCO',
    '106618': 'VILLAGE TECH',
    '106795': 'KOPJE SPARES',
    '010008': 'NICOZ',
    '106225': 'POTE',
    '105520': 'ALLIED INSURANCE',
    '105982': 'MISTY',
    '010736': 'CHICKEN HOTSPOT',
    '014508': 'CHICKEN HUT',
    '013745': 'ELEPHANT CHICKENS',
}

EXCLUDED_CIFS = {'012499', '011755', '014224', '013714', '014345', '004860'}
STMT_HIDE_PATTERNS = {'REVENUE', 'COMMISSION', 'COMM'}


# =============================================================================
# CORE HELPERS
# =============================================================================

def normalize_columns(df):
    """Strip and uppercase all column names.

    Guards against pandas converting date-formatted header cells to Timestamp /
    datetime objects (which stringify as '2026-04-01 00:00:00').  Any Timestamp
    column name is reformatted as DD-MMM-YY (e.g. '01-APR-26').
    """
    import datetime as _dt

    def _clean(c):
        if isinstance(c, pd.Timestamp):
            return c.strftime('%d-%b-%y').upper()
        if isinstance(c, (_dt.datetime, _dt.date)):
            return pd.Timestamp(c).strftime('%d-%b-%y').upper()
        return str(c).strip().upper()

    df.columns = [_clean(c) for c in df.columns]
    return df


def read_excel_with_preserved_headers(source, sheet_name=0):
    """Read an Excel sheet preserving date column header strings verbatim.

    pandas silently converts date-formatted header cells to Timestamp objects
    during a normal pd.read_excel call.  This helper reads the header row
    separately with ``dtype=object`` (which bypasses type inference) so that
    cells whose *value* is already a text string (e.g. "1 April 2026") come
    back as the original string rather than a Timestamp, while cells that ARE
    genuine date objects get formatted as DD-MMM-YY (no time component).

    Equivalent to ``normalize_columns(pd.read_excel(source, sheet_name))``,
    but without '2026-04-01 00:00:00' garbage in date column headers.
    """
    import datetime as _dt

    try:
        hdr_df = pd.read_excel(
            source, sheet_name=sheet_name,
            header=None, nrows=1, dtype=object,
        )
        raw_headers = list(hdr_df.iloc[0])
    except Exception:
        raw_headers = None

    df = pd.read_excel(source, sheet_name=sheet_name)

    if raw_headers is not None and len(raw_headers) == len(df.columns):
        new_cols = []
        for val in raw_headers:
            if val is None or (isinstance(val, float) and pd.isna(val)):
                new_cols.append('')
            elif isinstance(val, pd.Timestamp):
                new_cols.append(val.strftime('%d-%b-%y').upper())
            elif isinstance(val, (_dt.datetime, _dt.date)):
                new_cols.append(pd.Timestamp(val).strftime('%d-%b-%y').upper())
            else:
                new_cols.append(str(val).strip().upper())
        df.columns = new_cols
    else:
        df = normalize_columns(df)

    return df


def _parse_col_date(col_name):
    try:
        parts = re.split(r'[-/ ]', str(col_name).strip())
        if len(parts) >= 3:
            month_part = parts[1].upper()[:3]
            if month_part in MONTH_ABBR:
                day  = int(parts[0])
                year = int(parts[2])
                if year < 100:
                    year += 2000
                return pd.Timestamp(year=year, month=MONTH_ABBR[month_part], day=day)
            else:
                first, second, third = int(parts[0]), int(parts[1]), int(parts[2])
                if first > 31:
                    year, month, day = first, second, third
                else:
                    day, month, year = first, second, third
                if year < 100:
                    year += 2000
                return pd.Timestamp(year=year, month=month, day=day)
    except Exception:
        pass
    return pd.NaT


def extract_date_columns(df):
    return [col for col in df.columns if pd.notnull(_parse_col_date(col))]


def _to_numeric(series):
    return (
        series.astype(str)
        .str.replace(',', '', regex=False)
        .str.strip()
        .replace(['-', ' - ', '', 'nan', 'NaN', 'None'], '0')
        .pipe(pd.to_numeric, errors='coerce')
        .fillna(0)
    )


def _write_formatted_sheet(writer, df, sheet_name):
    workbook   = writer.book
    num_fmt    = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
    data_fmt   = workbook.add_format({'num_format': num_fmt})
    total_fmt  = workbook.add_format({'num_format': num_fmt, 'bold': True, 'top': 1, 'bottom': 6})
    label_fmt  = workbook.add_format({'bold': True, 'top': 1, 'bottom': 6})
    border_fmt = workbook.add_format({'top': 1, 'bottom': 6})
    header_fmt = workbook.add_format({'bold': True, 'bottom': 1})

    def _is_financial(col_name):
        return str(col_name).upper() == 'TOTAL' or pd.notnull(_parse_col_date(col_name))

    clean = df.copy()
    fin_idx = []
    for i, col in enumerate(clean.columns):
        if _is_financial(col):
            if not pd.api.types.is_numeric_dtype(clean[col]):
                clean[col] = _to_numeric(clean[col])
            else:
                clean[col] = clean[col].fillna(0)
            fin_idx.append(i)

    clean.to_excel(writer, index=False, sheet_name=sheet_name)
    ws = writer.sheets[sheet_name]

    ws.freeze_panes(1, 0)

    for ci, col_name in enumerate(clean.columns):
        ws.write(0, ci, str(col_name), header_fmt)

    for ci in fin_idx:
        for ri, val in enumerate(clean[clean.columns[ci]]):
            ws.write(ri + 1, ci, float(val), data_fmt)

    last_row = len(clean) + 1
    fin_set  = set(fin_idx)
    for ci in range(len(clean.columns)):
        if ci == 0:
            ws.write(last_row, ci, 'TOTAL', label_fmt)
        elif ci in fin_set:
            ws.write(last_row, ci, float(clean[clean.columns[ci]].sum()), total_fmt)
        else:
            ws.write(last_row, ci, '', border_fmt)


def _write_gainers_excel(gainers_df):
    buf = io.BytesIO()
    cols = ['MERCHANT NAME', 'PREVIOUSLY', 'CURRENTLY', 'VARIANCE']
    num_cols = [1, 2, 3]
    data = gainers_df[cols].reset_index(drop=True)

    with pd.ExcelWriter(buf, engine='xlsxwriter') as writer:
        workbook  = writer.book
        num_fmt   = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
        data_fmt  = workbook.add_format({'num_format': num_fmt})
        total_fmt = workbook.add_format({'num_format': num_fmt, 'bold': True, 'top': 1, 'bottom': 6})
        label_fmt = workbook.add_format({'bold': True, 'top': 1, 'bottom': 6})
        hdr_fmt   = workbook.add_format({'bold': True, 'bottom': 1})

        data.to_excel(writer, index=False, sheet_name='Gainers and Shakers')
        ws = writer.sheets['Gainers and Shakers']

        ws.freeze_panes(1, 0)

        for ci, hdr in enumerate(cols):
            ws.write(0, ci, hdr, hdr_fmt)

        for ri, row in data.iterrows():
            ws.write(ri + 1, 0, row['MERCHANT NAME'])
            for ci in num_cols:
                ws.write(ri + 1, ci, float(row[cols[ci]]), data_fmt)

        last_row = len(data) + 1
        ws.write(last_row, 0, 'TOTAL', label_fmt)
        for ci in num_cols:
            ws.write(last_row, ci, float(data[cols[ci]].sum()), total_fmt)

    return buf


def _build_report_sheets(rpt_df):
    sheets = {}

    sheets['INFRALINK TOLLS'] = rpt_df[
        rpt_df['MERCHANT NAME'].str.contains('INFRALINK', case=False, na=False)
    ]
    sheets['CHAMPIONS - NBS'] = rpt_df[
        (rpt_df['CIF'] == '012499') &
        (rpt_df['TID'].str.contains('NBSP', case=False, na=False))
    ]
    sheets['CHAMPIONS INSURANCE'] = rpt_df[
        (rpt_df['CIF'] == '012499') &
        (rpt_df['TID'].str.contains('NBSX', case=False, na=False))
    ]
    sheets['ZINARA LICENCING'] = rpt_df[
        (rpt_df['CIF'] == '011755') &
        (rpt_df['MERCHANT NAME'].str.contains('LICENCING', case=False, na=False))
    ]
    sheets['BANCASSURANCE (NBS)'] = rpt_df[rpt_df['TID'].isin(BA_TIDS)]
    sheets['ZINARA TOLLGATES']    = rpt_df[rpt_df['TID'].isin(ZT_TIDS)]
    sheets['NSSA'] = rpt_df[
        rpt_df['MERCHANT NAME'].str.contains('NSSA', case=False, na=False)
    ]
    sheets['CHICKEN SLICE'] = rpt_df[rpt_df['CIF'].isin(['013714', '014345'])]

    named_cifs = set(
        rpt_df[rpt_df['MERCHANT NAME'].str.contains('INFRALINK', case=False, na=False)]['CIF'].tolist() +
        rpt_df[rpt_df['MERCHANT NAME'].str.contains('NSSA',      case=False, na=False)]['CIF'].tolist()
    )
    excluded = EXCLUDED_CIFS | named_cifs
    counts   = rpt_df['CIF'].value_counts()
    eligible = counts[(counts > 7) & (~counts.index.isin(excluded))].index

    used_names = set(sheets.keys())
    for cif in eligible:
        # Skip zero/null CIFs — these are placeholder rows where the CIF column
        # held 0, 0.0, or NaN; after zfill they become '000000' or contain 'nan'.
        _cif_str = str(cif)
        if not _cif_str or _cif_str.lstrip('0') == '' or 'nan' in _cif_str.lower():
            continue

        cif_data = rpt_df[rpt_df['CIF'] == cif]
        if cif in CIF_SHEET_NAMES:
            _derived = CIF_SHEET_NAMES[cif]
        else:
            # Find first genuinely valid merchant name for this CIF.
            # Exclude empty strings, 'nan'/'none', and zero-valued numerics
            # ('0', '0.0') that pandas produces when a cell held the integer 0.
            _names = (
                cif_data['MERCHANT NAME']
                .dropna()
                .astype(str)
                .str.strip()
            )
            _invalid = {'', 'nan', 'none', '0', '0.0', '0.00'}
            _names = _names[~_names.str.lower().isin(_invalid)]
            if _names.empty:
                continue  # No real merchant name — omit this CIF entirely
            _derived = _names.iloc[0].split()[0]
        sheet_name = _derived[:31]
        if sheet_name in used_names:
            sheet_name = (sheet_name[:27] + '_' + cif[-3:])[:31]
        used_names.add(sheet_name)
        sheets[sheet_name] = cif_data

    return sheets


def _build_gainers_shakers(report_sheets, date_map, latest_col_name, threshold):
    all_date_cols = [date_map[d] for d in sorted(date_map.keys())]
    if latest_col_name in all_date_cols:
        idx = all_date_cols.index(latest_col_name)
        prior_cols = all_date_cols[:idx]
    else:
        prior_cols = []

    records = []
    for sheet_name, data in report_sheets.items():
        if str(sheet_name).strip().lower() in ('nan', 'none', ''):
            continue
        if latest_col_name not in data.columns:
            continue

        col_curr = _to_numeric(data[latest_col_name]).sum()
        if col_curr == 0:
            continue

        col_prev = None
        for c in reversed(prior_cols):
            if c in data.columns:
                candidate = _to_numeric(data[c]).sum()
                if candidate != 0:
                    col_prev = candidate
                    break

        if col_prev is None:
            continue

        variance = col_curr - col_prev
        if variance >= threshold or variance <= -threshold:
            records.append({
                'MERCHANT NAME': sheet_name,
                'PREVIOUSLY':    col_prev,
                'CURRENTLY':     col_curr,
                'VARIANCE':      variance,
            })

    if not records:
        return pd.DataFrame(columns=['MERCHANT NAME', 'PREVIOUSLY', 'CURRENTLY', 'VARIANCE'])

    return (
        pd.DataFrame(records)
        .sort_values('VARIANCE', ascending=False)
        .reset_index(drop=True)
    )


# =============================================================================
# DASHBOARD DATA BUILDERS
# =============================================================================

def resolve_dates(df, date_columns):
    date_map = {}
    for col in date_columns:
        parsed = _parse_col_date(col)
        if pd.notnull(parsed):
            date_map[parsed] = col

    sorted_dates = sorted(date_map.keys())

    latest = sorted_dates[-1]
    for d in reversed(sorted_dates):
        if df[date_map[d]].sum() > 0:
            latest = d
            break

    idx  = sorted_dates.index(latest)
    prev = sorted_dates[idx - 1] if idx > 0 else None
    return date_map, sorted_dates, latest, date_map[latest], prev


def apply_cif_groups(summary):
    for group in CIF_GROUPS:
        members = summary[summary['CIF'].isin(group['cifs'])]
        if members.empty:
            continue

        combined_revenue = members['Revenue'].sum()
        combined_dc      = members['Daily Change'].sum()
        combined_prev    = (members['Revenue'] - members['Daily Change']).sum()
        combined_growth  = (combined_dc / combined_prev * 100) if combined_prev > 0 else 0.0
        combined_active  = members['Active_Terminals'].sum()
        combined_total   = members['Total_Terminals'].sum()
        combined_usage   = round(combined_active / combined_total * 100, 1) if combined_total > 0 else 0.0

        primary  = members[members['CIF'] == group['primary_cif']]
        biz_unit = primary['BizUnit'].iloc[0] if not primary.empty else members['BizUnit'].iloc[0]

        summary = summary[~summary['CIF'].isin(group['cifs'])]
        summary = pd.concat([summary, pd.DataFrame([{
            'CIF':              group['primary_cif'],
            'Merchant':         group['merchant'],
            'BizUnit':          biz_unit,
            'Revenue':          combined_revenue,
            'Active_Terminals': combined_active,
            'Total_Terminals':  combined_total,
            'Daily Change':     combined_dc,
            'Growth %':         combined_growth,
            'Usage Ratio':      combined_usage,
        }])], ignore_index=True)

    return summary


def build_top15_data(df, date_map, latest_date_obj, latest_col_name, prev_date_obj):
    """Return (metrics_dict, top15_rows_list) for template rendering."""
    daily_data = df[df[latest_col_name] > 0].copy()
    if daily_data.empty:
        return None, []

    summary = daily_data.groupby('CIF').agg(
        Merchant=('MERCHANT NAME', 'first'),
        BizUnit=('BUSINESS UNIT', 'first'),
        Revenue=(latest_col_name, 'sum'),
        Active_Terminals=('TID', 'nunique'),
    ).reset_index()

    summary['Merchant'] = summary['Merchant'].astype(str).str.split().str[0]
    summary['Merchant'] = summary.apply(
        lambda r: CIF_NAME_OVERRIDES.get(r['CIF'], r['Merchant']), axis=1
    )
    summary['BizUnit'] = summary['BizUnit'].astype(str).str.split().str[0]
    summary['BizUnit'] = summary.apply(
        lambda r: CIF_BIZUNIT_OVERRIDES.get(r['CIF'], r['BizUnit']), axis=1
    )

    if prev_date_obj:
        prev_col = date_map[prev_date_obj]
        prev_rev = df.groupby('CIF')[prev_col].sum().reset_index(name='Prev_Revenue')
        summary  = summary.merge(prev_rev, on='CIF', how='left').fillna(0)
        summary['Daily Change'] = summary['Revenue'] - summary['Prev_Revenue']
        summary['Growth %'] = summary.apply(
            lambda x: (x['Daily Change'] / x['Prev_Revenue'] * 100) if x['Prev_Revenue'] > 0 else 0,
            axis=1,
        )
    else:
        summary['Daily Change'] = 0
        summary['Growth %']     = 0.0

    total_tids = df.groupby('CIF')['TID'].nunique().reset_index(name='Total_Terminals')
    summary = summary.merge(total_tids, on='CIF', how='left')
    summary['Usage Ratio'] = (summary['Active_Terminals'] / summary['Total_Terminals'] * 100).round(1)
    summary = apply_cif_groups(summary)

    top_15 = summary.sort_values('Revenue', ascending=False).head(15)

    _cif_blank = (
        daily_data['CIF'].isna() |
        daily_data['CIF'].astype(str).str.strip().isin(['', 'nan', 'NaN', 'None'])
    )
    _valid_tid = ~daily_data['TID'].astype(str).str.strip().str.upper().isin(
        ['', 'TOTAL', 'GRAND TOTAL', 'NAN', 'NONE']
    )
    orphan_rev    = daily_data.loc[_cif_blank & _valid_tid, latest_col_name].sum()
    total_day_rev = round(summary['Revenue'].sum() + orphan_rev, 2)

    metrics = {
        'total_revenue':     f'{total_day_rev:,.0f}',
        'mtd_revenue':       '—',
        'active_merchants':  summary['CIF'].nunique(),
        'active_terminals':  int(summary['Active_Terminals'].sum()),
        'top_merchant':      top_15.iloc[0]['Merchant'] if not top_15.empty else '—',
        'display_date':      latest_date_obj.strftime('%d %b %Y'),
    }

    rows = []
    for _, row in top_15.iterrows():
        daily_ch = row['Daily Change']
        rows.append({
            'cif':       row['CIF'],
            'merchant':  row['Merchant'],
            'bizunit':   row['BizUnit'],
            'revenue':   f"{row['Revenue']:,.0f}",
            'daily_ch':  f"{'+'if daily_ch>=0 else ''}{daily_ch:,.0f}",
            'ch_pos':    daily_ch >= 0,
            'growth':    f"{row['Growth %']:.1f}%",
            'usage':     f"{row['Usage Ratio']:.1f}%",
            'usage_low': row['Usage Ratio'] < 50,
        })

    return metrics, rows


def build_plotly_bar_div(labels, values, title, x_label):
    """Return an HTML div string for a horizontal bar chart (Plotly)."""
    n = len(labels)
    colours = [
        f"hsl({200 + int(i * 40 / max(n - 1, 1))}, 70%, {55 - int(i * 15 / max(n - 1, 1))}%)"
        for i in range(n)
    ]
    fig = go.Figure(go.Bar(
        x=values,
        y=labels,
        orientation='h',
        marker=dict(color=colours, line=dict(color='rgba(255,255,255,0.15)', width=0.8)),
        text=[f'{v:,.0f}' for v in values],
        textposition='outside',
        textfont=dict(color='#FFD700', size=12, family='Arial Black'),
        hovertemplate='<b>%{y}</b><br>' + x_label + ': %{x:,.0f}<extra></extra>',
    ))
    fig.update_layout(
        title=dict(text=title, font=dict(color='#e8f4fd', size=16, family='Arial Black'), x=0.01),
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(255,255,255,0.04)',
        xaxis=dict(
            title=x_label,
            title_font=dict(color='#a8d8f0', size=13),
            tickfont=dict(color='#a8d8f0'),
            tickformat=',.0f',
            gridcolor='rgba(255,255,255,0.08)',
            showgrid=True,
            zeroline=False,
        ),
        yaxis=dict(tickfont=dict(color='#e0e8f0', size=12), autorange='reversed', showgrid=False),
        margin=dict(l=10, r=80, t=50, b=40),
        height=460,
        showlegend=False,
    )
    return pyo.plot(fig, output_type='div', include_plotlyjs=False)


def compute_mtd_from_total_row(df_raw, date_columns):
    """
    Read the MTD from the TOTAL/GRAND TOTAL row.
    Prefers the dedicated TOTAL meta-column (which respects the report's
    formula range, correctly excluding rows added outside that range).
    Falls back to summing date columns if TOTAL column is absent or zero.
    """
    cols = [c for c in date_columns if c in df_raw.columns]
    if not cols:
        return 0.0

    total_mask = (
        df_raw['MERCHANT NAME'].astype(str).str.upper().str.strip().isin(['TOTAL', 'GRAND TOTAL']) |
        df_raw['TID'].astype(str).str.upper().str.strip().isin(['TOTAL', 'GRAND TOTAL'])
    )
    total_rows = df_raw[total_mask]

    if not total_rows.empty:
        total_row = total_rows.iloc[-1]

        # Prefer the TOTAL meta-column: it reflects the workbook's own SUM
        # formula range, which intentionally excludes any rows appended
        # outside the original formula range (e.g. newly inserted terminals).
        if 'TOTAL' in df_raw.columns:
            val = _to_numeric(pd.Series([total_row['TOTAL']])).iloc[0]
            if val > 0:
                return round(float(val), 2)

        # Fallback: sum the date columns in the TOTAL row
        mtd = sum(
            _to_numeric(pd.Series([total_row[c]])).iloc[0]
            for c in cols
        )
    else:
        mtd = df_raw[cols].apply(_to_numeric).sum().sum()

    return round(float(mtd), 2)


def compute_daily_from_total_row(df_raw, col_name):
    """Read the daily revenue for a specific date column from the TOTAL row."""
    if col_name not in df_raw.columns:
        return 0.0
    total_mask = (
        df_raw['MERCHANT NAME'].astype(str).str.upper().str.strip().isin(['TOTAL', 'GRAND TOTAL']) |
        df_raw['TID'].astype(str).str.upper().str.strip().isin(['TOTAL', 'GRAND TOTAL'])
    )
    total_rows = df_raw[total_mask]
    if not total_rows.empty:
        val = _to_numeric(pd.Series([total_rows.iloc[-1][col_name]])).iloc[0]
        if val > 0:
            return round(float(val), 2)
    # Fallback: sum individual rows excluding total rows
    data_rows = df_raw[~total_mask]
    return round(float(_to_numeric(data_rows[col_name]).sum()), 2)


def build_performance_report_bytes(df_raw):
    """Build the daily performance Excel report and return bytes."""
    rpt_df = df_raw.copy()
    rpt_df['CIF']           = rpt_df['CIF'].astype(str).str.strip().str.split('.').str[0].str.zfill(6)
    rpt_df['TID']           = rpt_df['TID'].astype(str).str.strip()
    rpt_df['MERCHANT NAME'] = rpt_df['MERCHANT NAME'].astype(str).str.strip()

    report_sheets = _build_report_sheets(rpt_df)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='xlsxwriter') as writer:
        for sname, data in report_sheets.items():
            if not data.empty:
                _write_formatted_sheet(writer, data, sname[:31])
    return buf.getvalue(), report_sheets


def build_top25_data(df, fixed_cols, date_columns, start_d, end_d):
    """Return (top25_df, total_revenue) or (None, error_msg)."""
    txn_df    = df.melt(id_vars=list(fixed_cols), value_vars=date_columns, var_name='D_COL', value_name='AMT')
    txn_df['DATE'] = txn_df['D_COL'].apply(_parse_col_date)
    period_df = txn_df[(txn_df['DATE'].dt.date >= start_d) & (txn_df['DATE'].dt.date <= end_d)]

    if period_df.empty:
        return None, f'No transactions found between {start_d} and {end_d}.'

    summary_p = (
        period_df.groupby('CIF')
        .agg(
            Merchant_Name=('MERCHANT NAME', 'first'),
            Business_Unit=('BUSINESS UNIT', 'first'),
            Revenue=('AMT', 'sum'),
        )
        .reset_index()
        .rename(columns={'Merchant_Name': 'MERCHANT NAME', 'Business_Unit': 'BUSINESS UNIT'})
    )
    total_r = summary_p['Revenue'].sum()
    summary_p['Revenue Percentage'] = summary_p['Revenue'] / total_r * 100

    top_25 = summary_p.sort_values('Revenue', ascending=False).head(25).copy()
    top_25['Revenue Percentage'] = top_25['Revenue Percentage'].round(0).astype(int).astype(str) + '%'
    return top_25, total_r


def build_top25_excel_bytes(top_25):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        top_25[['CIF', 'MERCHANT NAME', 'Revenue', 'BUSINESS UNIT', 'Revenue Percentage']].to_excel(
            writer, index=False, sheet_name='Top 25'
        )
        writer.sheets['Top 25'].freeze_panes(1, 0)
    return output.getvalue()


def find_idle_terminals(df_zwg, df_usd, zwg_date_cols, usd_date_cols, from_date, to_date):
    """Return (total_idle_df, idle_zwg_df, idle_usd_df, common_sfx) or raise ValueError."""
    def _cols_in_range(date_cols, from_d, to_d):
        return [
            c for c in date_cols
            if pd.notnull(_parse_col_date(c)) and from_d <= _parse_col_date(c).date() <= to_d
        ]

    zwg_period = _cols_in_range(zwg_date_cols, from_date, to_date)
    usd_period = _cols_in_range(usd_date_cols, from_date, to_date)

    if not zwg_period and not usd_period:
        raise ValueError(f'No date columns found within the selected period.')

    def _find_idle(frame, cols):
        if not cols:
            return pd.DataFrame()
        frame = frame.copy()
        frame[cols] = frame[cols].apply(_to_numeric)
        frame['_PT'] = frame[cols].sum(axis=1)
        return frame[frame['_PT'] == 0].drop(columns=['_PT'])

    idle_zwg = _find_idle(df_zwg, zwg_period)
    idle_usd = _find_idle(df_usd, usd_period)

    if idle_zwg.empty and idle_usd.empty:
        raise ValueError('No idle terminals found for the selected period.')

    idle_zwg['_Sfx'] = idle_zwg['TID'].str[-4:]
    usd_sfx    = set(idle_usd['TID'].str[-4:]) if not idle_usd.empty else set()
    common_sfx = set(idle_zwg['_Sfx']).intersection(usd_sfx)
    total_idle = idle_zwg[idle_zwg['_Sfx'].isin(common_sfx)].drop(columns=['_Sfx']).copy()
    idle_zwg   = idle_zwg.drop(columns=['_Sfx'])

    if total_idle.empty:
        raise ValueError('No common idle terminals found between ZWG and USD files for this period.')

    return total_idle, idle_zwg, idle_usd, common_sfx


def build_merchant_revenue_chart(daily_zwg, daily_usd, merchant_name, date_label):
    """Daily ZWG + USD grouped bar chart for a single merchant."""
    all_dates = sorted(set(list(daily_zwg.keys()) + list(daily_usd.keys())))
    if not all_dates:
        return ''
    date_strs = [d.strftime('%d %b') for d in all_dates]
    fig = go.Figure()
    if any(v > 0 for v in daily_zwg.values()):
        fig.add_trace(go.Bar(
            x=date_strs, y=[daily_zwg.get(d, 0) for d in all_dates],
            name='ZWG', marker_color='#FFD700',
            hovertemplate='<b>%{x}</b><br>ZWG: %{y:,.0f}<extra></extra>',
        ))
    if any(v > 0 for v in daily_usd.values()):
        fig.add_trace(go.Bar(
            x=date_strs, y=[daily_usd.get(d, 0) for d in all_dates],
            name='USD', marker_color='#4fc97e',
            hovertemplate='<b>%{x}</b><br>USD: %{y:,.2f}<extra></extra>',
        ))
    fig.update_layout(
        title=dict(
            text=f'{merchant_name} — Daily Revenue | {date_label}',
            font=dict(color='#e8f4fd', size=14, family='Arial Black'), x=0.01,
        ),
        paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(255,255,255,0.04)',
        xaxis=dict(tickfont=dict(color='#a8d8f0', size=10), gridcolor='rgba(255,255,255,0.08)'),
        yaxis=dict(tickfont=dict(color='#a8d8f0'), tickformat=',.0f', gridcolor='rgba(255,255,255,0.08)'),
        legend=dict(font=dict(color='#e0e8f0', size=12), bgcolor='rgba(0,0,0,0)'),
        barmode='group', margin=dict(l=10, r=20, t=50, b=40), height=360,
    )
    return pyo.plot(fig, output_type='div', include_plotlyjs=False)


def merchant_period_excel_bytes(filt_zwg, filt_usd, zwg_period_cols, usd_period_cols):
    """Period-filtered performance report: ZWG and USD sheets with only the chosen date range."""
    fixed = ['TID', 'CIF', 'MERCHANT NAME', 'BUSINESS UNIT']

    def _prepare(frame, period_cols):
        if frame is None or frame.empty:
            return pd.DataFrame()
        avail_fixed  = [c for c in fixed if c in frame.columns]
        avail_period = [c for c in period_cols if c in frame.columns]
        if not avail_period:
            return pd.DataFrame()
        df = frame[avail_fixed + avail_period].copy()
        for col in avail_period:
            df[col] = _to_numeric(df[col])
        # The all-sheets concat gives one row per TID per month-sheet.
        # Collapse duplicates: sum period columns, keep first value for text.
        if 'TID' in df.columns:
            agg = {c: 'first' for c in avail_fixed if c != 'TID'}
            agg.update({c: 'sum' for c in avail_period})
            df = df.groupby('TID', as_index=False).agg(agg)
            df = df[[c for c in avail_fixed + avail_period if c in df.columns]]
        df['TOTAL'] = df[avail_period].sum(axis=1)
        return df

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='xlsxwriter') as writer:
        zwg_data = _prepare(filt_zwg, zwg_period_cols)
        usd_data = _prepare(filt_usd, usd_period_cols)
        wrote    = False
        if not zwg_data.empty:
            _write_formatted_sheet(writer, zwg_data, 'ZWG')
            wrote = True
        if not usd_data.empty:
            _write_formatted_sheet(writer, usd_data, 'USD')
            wrote = True
        if not wrote:
            pd.DataFrame({'Note': ['No data found for the selected period']}).to_excel(
                writer, index=False, sheet_name='No Data'
            )
    return buf.getvalue()


def idle_excel_bytes(df):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Idle Terminals')
        writer.sheets['Idle Terminals'].freeze_panes(1, 0)
    return buf.getvalue()


def _recalc_total_col(frame):
    """
    Return a copy of *frame* where the TOTAL column is recomputed from scratch
    as the row-wise sum of every date column (beginning of month → current date).

    This corrects TIDs whose TOTAL cell was missing, zero, or came from an
    Excel formula whose range did not cover all date columns (a common cause of
    under-reported POS statement totals).

    If the frame has no TOTAL column one is inserted immediately after the last
    date column.  Date columns that are still object-typed are coerced to numeric
    before summing so that text dashes ("-") and blanks count as zero.
    """
    df = frame.copy()

    # Identify all date columns present after STMT_HIDE_PATTERNS columns were dropped
    date_cols = [c for c in df.columns if pd.notnull(_parse_col_date(c))]
    if not date_cols:
        return df  # Nothing to sum — leave frame unchanged

    # Coerce every date column to numeric (blanks / dashes → 0)
    for col in date_cols:
        if not pd.api.types.is_numeric_dtype(df[col]):
            df[col] = _to_numeric(df[col])
        else:
            df[col] = df[col].fillna(0)

    # Recompute TOTAL as row-wise sum of ALL date columns
    recalculated = df[date_cols].sum(axis=1)

    if 'TOTAL' in df.columns:
        df['TOTAL'] = recalculated
    else:
        # Insert TOTAL column right after the last date column
        last_date_pos = max(df.columns.get_loc(c) for c in date_cols)
        df.insert(last_date_pos + 1, 'TOTAL', recalculated)

    return df


def _clean_stmt(frame):
    """Drop hidden columns (REVENUE, COMMISSION, COMM) then fix the TOTAL column."""
    df = frame.drop(
        columns=[c for c in frame.columns if any(p in c.upper() for p in STMT_HIDE_PATTERNS)],
        errors='ignore',
    )
    return _recalc_total_col(df)


def pos_statement_bytes(filt_zwg, filt_usd):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='xlsxwriter') as writer:
        _write_formatted_sheet(writer, _clean_stmt(filt_zwg), 'ZWG')
        _write_formatted_sheet(writer, _clean_stmt(filt_usd), 'USD')
    return buf.getvalue()


def champions_zinara_bytes(filt_zwg, filt_usd):
    zinara_zwg = filt_zwg[filt_zwg['TID'].str.contains('NBSP', case=False, na=False)]
    zinara_usd = filt_usd[filt_usd['TID'].str.contains('NBSP', case=False, na=False)]
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='xlsxwriter') as writer:
        _write_formatted_sheet(writer, _clean_stmt(zinara_zwg), 'ZWG')
        _write_formatted_sheet(writer, _clean_stmt(zinara_usd), 'USD')
    return buf.getvalue()


def champions_insurance_bytes(filt_zwg, filt_usd):
    insur_zwg = filt_zwg[filt_zwg['TID'].str.contains('NBSX', case=False, na=False)]
    insur_usd = filt_usd[filt_usd['TID'].str.contains('NBSU', case=False, na=False)]
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='xlsxwriter') as writer:
        _write_formatted_sheet(writer, _clean_stmt(insur_zwg), 'ZWG')
        _write_formatted_sheet(writer, _clean_stmt(insur_usd), 'USD')
    return buf.getvalue()


# =============================================================================
# AUTOMATIC REPORTS GENERATION
# =============================================================================

_B02_KNOWN_TYPES = {
    'goods and services',
    'goods and services (rev)',
    'goods services with cash back',
    'cash withdrawal',
    'cash withdrawal (rev)',
    'balance inquiry',
    'balance inquiry (rev)',
    'deposit',
    'deposit (rev)',
    'available funds inquiry',
}


def _read_b02_df(raw_bytes, fname):
    """
    Read a B02 Excel file into a DataFrame, trying xlrd for .xls and openpyxl
    for .xlsx, with a cross-engine fallback for mislabelled files.
    Returns a DataFrame or raises the last exception encountered.
    """
    is_xls = fname.lower().endswith('.xls')
    errors = []
    for engine in (['xlrd', 'openpyxl'] if is_xls else ['openpyxl', 'xlrd']):
        try:
            return pd.read_excel(io.BytesIO(raw_bytes), header=None, engine=engine)
        except Exception as e:
            errors.append(e)
    raise errors[0]


def _col_is_blank(series):
    """True for each element that is NaN or an empty/whitespace string."""
    return series.isna() | (series.fillna('').astype(str).str.strip() == '')


def validate_b02_file(raw_bytes, fname):
    """
    Validate that raw_bytes is a genuine B02 Balancing Terminal Detail file.
    Returns None when valid, or a human-readable error string when not.

    Checks (in order):
      1. Readable as .xls / .xlsx
      2. Has at least 11 columns (B02 has 17)
      3. Contains TID header rows  (col0 starts "NB", col1 is blank)
      4. Transaction rows have datetime values in col0
      5. Known B02 transaction types present in col3
      6. Currency column (col10) contains only ZWG / USD
      7. "Net Surcharge" per-terminal footer rows are present
    """
    # 1 ── Readable as Excel ────────────────────────────────────────────────────
    try:
        df = _read_b02_df(raw_bytes, fname)
    except Exception as e:
        return (
            f'"{fname}" could not be read as an Excel file '
            f'({type(e).__name__}: {e}). '
            f'B02 reports must be .xls or .xlsx. '
            f'Please upload an original NBS B02 Balancing Terminal Detail file.'
        )

    # 2 ── Minimum columns ─────────────────────────────────────────────────────
    if df.empty or df.shape[1] < 11:
        return (
            f'"{fname}" does not appear to be a B02 report — '
            f'expected at least 11 columns but found {df.shape[1]}. '
            f'Please upload an original NBS B02 Balancing Terminal Detail file.'
        )

    col0_raw = df.iloc[:, 0]
    col0_str = col0_raw.fillna('').astype(str).str.strip()
    col1_raw = df.iloc[:, 1]
    col3_str = df.iloc[:, 3].fillna('').astype(str).str.strip().str.lower()

    # 3 ── TID header rows ─────────────────────────────────────────────────────
    # col1 must be blank (NaN or empty string — xlrd can return either)
    tid_mask = col0_str.str.upper().str.startswith('NB') & _col_is_blank(col1_raw)
    if not tid_mask.any():
        return (
            f'"{fname}" does not appear to be a B02 Balancing Terminal Detail report. '
            f'No terminal (TID) header rows were found. '
            f'In a genuine B02 file each terminal section begins with a row whose '
            f'first column contains the terminal ID (e.g. "NB020501") and all '
            f'other columns are blank. '
            f'Please upload an original NBS B02 Balancing Terminal Detail file.'
        )

    # 4 ── Transaction rows have datetime values in col0 ───────────────────────
    non_tid = ~tid_mask & col0_str.ne('')
    dt_count = pd.to_datetime(col0_raw[non_tid], errors='coerce').notna().sum()
    if dt_count == 0:
        return (
            f'"{fname}" does not appear to be a B02 report. '
            f'Transaction rows must have date/time values in the first column '
            f'(e.g. "2026-05-15 12:07:45"), but none were found. '
            f'Please upload an original NBS B02 Balancing Terminal Detail file.'
        )

    # 5 ── Known B02 transaction types in col3 ─────────────────────────────────
    if not col3_str.isin(_B02_KNOWN_TYPES).any():
        return (
            f'"{fname}" does not appear to be a B02 report. '
            f'Expected transaction types such as "Goods and services" or '
            f'"Cash withdrawal" in column 4, but none were recognised. '
            f'Please upload an original NBS B02 Balancing Terminal Detail file.'
        )

    # 6 ── Currency column contains only ZWG / USD ─────────────────────────────
    currencies = (
        df.iloc[:, 10].dropna().astype(str).str.strip().str.upper()
    )
    currencies = currencies[
        currencies.ne('') & ~currencies.str.lower().str.startswith('net')
    ]
    if not currencies.empty:
        unknown = set(currencies.unique()) - {'ZWG', 'USD'}
        if unknown and unknown == set(currencies.unique()):
            return (
                f'"{fname}" does not appear to be an NBS B02 report. '
                f'Expected currencies ZWG and/or USD in column 11 but found: '
                f'{", ".join(sorted(unknown))}. '
                f'Please upload a file generated by the NBS switch.'
            )

    # 7 ── "Net Surcharge" per-terminal footer rows ────────────────────────────
    if not col0_str.str.lower().str.startswith('net surcharge').any():
        return (
            f'"{fname}" does not appear to be a B02 Balancing Terminal Detail report. '
            f'The per-terminal summary rows (beginning with "Net Surcharge") were '
            f'not found. '
            f'Please upload an original NBS B02 Balancing Terminal Detail file.'
        )

    return None  # ✓ valid


def parse_b02_files(file_list):
    """
    Parse one or more validated B02 raw switch files.
    Returns {(tid, date_obj, currency_upper): total_amount}

    B02 structure (confirmed from file analysis):
      TID header rows:  col0 = terminal ID starting "NB", col1 = NaN
      Transaction rows: col0 = transaction datetime, col1 = sequence number,
                        col3 = transaction type, col5 = status,
                        col7 = gross transaction amount, col10 = currency (ZWG or USD)
      Per-TID footer:   col0 = "Net Surcharge"  (skipped)

    Only "Goods and services" (and "Goods services with cash back") transactions
    with status "Approved or completed successfully" are counted.

    Business date: ALL transactions in a B02 file belong to one business day
    regardless of the actual datetime in col0 (which can contain stale/overnight
    entries from previous or next calendar days). Business date = YYYYMMDD from
    the filename minus 1 day.  Col0 is NOT used for date assignment.

    Amount column: col7 (gross settled amount, before any surcharge deduction).
    Col9 is the net-after-surcharge figure which understates the merchant total.
    """
    totals = {}

    for f in file_list:
        fname = getattr(f, 'name', '') or ''
        try:
            raw = f.read()
        except Exception:
            continue

        try:
            df = _read_b02_df(raw, fname)
        except Exception:
            continue

        # Business date = YYYYMMDD from filename minus 1 day.
        # All transactions in this file belong to this single business date.
        business_date = None
        m = re.search(r'(\d{8})', fname)
        if m:
            try:
                business_date = (pd.Timestamp(m.group(1)) - pd.Timedelta(days=1)).date()
            except Exception:
                pass
        if business_date is None:
            continue  # cannot determine business date — skip file

        current_tid = None
        for _, row in df.iterrows():
            if len(row) < 11:
                continue
            col0 = row.iloc[0]
            col1 = row.iloc[1]
            col0_str = str(col0).strip() if pd.notna(col0) else ''

            # TID header row: col0 = "NB...", col1 = blank (NaN or empty string)
            if col0_str.upper().startswith('NB') and (
                pd.isna(col1) or str(col1).strip() == ''
            ):
                current_tid = col0_str
                continue

            if current_tid is None:
                continue

            # Skip footer and page-text rows
            col0_lower = col0_str.lower()
            if col0_lower.startswith('net') or col0_lower.startswith('this report'):
                continue

            col3 = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ''
            col5 = str(row.iloc[5]).strip() if pd.notna(row.iloc[5]) else ''

            # Only "Goods and services" (plain) — exclude "Goods services with cash back"
            # because for cash-back transactions the gross amount includes cash dispensed
            # to the customer, which is not merchant goods revenue.
            col3_lower = col3.lower()
            if 'goods' not in col3_lower or 'cash back' in col3_lower:
                continue
            if 'approved or completed' not in col5.lower():
                continue

            # col7 = gross transaction amount (before surcharge deduction)
            try:
                amount = float(row.iloc[7])
            except (ValueError, TypeError):
                continue
            if amount <= 0:
                continue

            currency = str(row.iloc[10]).strip().upper() if pd.notna(row.iloc[10]) else ''
            if currency not in ('ZWG', 'USD'):
                continue

            key = (current_tid, business_date, currency)
            totals[key] = totals.get(key, 0.0) + amount

    return totals


def update_merchant_report(report_bytes, b02_totals, currency):
    """
    Update an Excel merchant report in-place using openpyxl.

    For each matching sheet (where a date column matches a B02 transaction date):
      - existing TID rows in B02: add the amount to the date cell
      - existing TID rows NOT in B02: set 0 (shown as '-') for that date column
      - new TIDs: insert a row before TOTAL (0 for past dates, amount for B02
        date, blank for future dates); borders cleared, accounting format applied
    Recalculates the per-row TOTAL meta-column and the TOTAL summary row.

    Returns (updated_bytes, stats_dict) with keys 'updated' and 'added'.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Border, Side, Font

    _NO_BORDER = Border()   # blank border object used to clear inherited styles
    _FALLBACK_FMT = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'

    currency_upper = currency.upper()

    # Filter to matching currency: {(tid, date_obj): amount}
    relevant = {}
    for (tid, d, cur), amt in b02_totals.items():
        if cur.upper() == currency_upper:
            key = (tid, d)
            relevant[key] = relevant.get(key, 0.0) + amt

    if not relevant:
        return report_bytes, {'updated': 0, 'added': 0}

    wb = load_workbook(io.BytesIO(report_bytes), keep_links=False)
    stats = {'updated': 0, 'added': 0}
    import datetime as _dt
    _all_date_totals = {}   # accumulated per-date totals across all processed sheets

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue

        # Read header row
        # openpyxl returns datetime objects for date-formatted cells; convert
        # them to a clean DD-MMM-YY string so header_map keys are consistent.
        import datetime as _dt_hdr
        header_map = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(1, c).value
            if val is not None:
                if isinstance(val, (_dt_hdr.datetime, _dt_hdr.date)):
                    _hk = pd.Timestamp(val).strftime('%d-%b-%y').upper()
                else:
                    _hk = str(val).strip().upper()
                header_map[_hk] = c

        if 'TID' not in header_map:
            continue

        tid_col        = header_map['TID']
        total_meta_col = header_map.get('TOTAL')
        revenue_col    = header_map.get('REVENUE')

        # Build date -> column map from header row
        date_col_map = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(1, c).value
            if val is not None:
                if isinstance(val, (_dt_hdr.datetime, _dt_hdr.date)):
                    _parsed = pd.Timestamp(val)
                else:
                    _parsed = _parse_col_date(str(val))
                if pd.notna(_parsed):
                    date_col_map[_parsed.date()] = c

        if not date_col_map:
            continue

        # Only process sheets that contain at least one B02 date
        b02_dates = {d for (_, d) in relevant}
        matching_b02_dates = b02_dates & set(date_col_map.keys())
        if not matching_b02_dates:
            continue

        # ── Detect the accounting number format used by existing data cells ──
        # Scan the first few data rows/date columns for a non-General format.
        num_fmt = _FALLBACK_FMT
        sample_cols = list(date_col_map.values())[:5]
        outer_break = False
        for r in range(2, min(ws.max_row, 30)):
            for col in sample_cols:
                fmt = ws.cell(r, col).number_format
                if fmt and fmt.upper() not in ('GENERAL', '@', ''):
                    num_fmt = fmt
                    outer_break = True
                    break
            if outer_break:
                break

        # Find TOTAL row (scan from bottom for TID col = "TOTAL")
        total_row = None
        for r in range(ws.max_row, 1, -1):
            val = ws.cell(r, tid_col).value
            if val is not None and str(val).strip().upper() == 'TOTAL':
                total_row = r
                break
        if total_row is None:
            total_row = ws.max_row

        # Save TOTAL row borders so we can restore them after insert_rows() calls.
        # Save ALL columns (including cells with no border = Border()) so that
        # any border accidentally copied by insert_rows() is also cleared.
        from copy import copy as _copy_border
        _saved_total_borders = {}
        for _sc in range(1, ws.max_column + 1):
            _orig = ws.cell(total_row, _sc)
            _saved_total_borders[_sc] = _copy_border(_orig.border) if _orig.border else Border()

        # Find a REVENUE formula from an existing data row to use as a template
        # when inserting new TID rows (so REVENUE is populated for new entries).
        _rev_formula_template = None
        if revenue_col:
            for _scan_r in range(2, min(total_row, 100)):
                _rv_val = ws.cell(_scan_r, revenue_col).value
                if isinstance(_rv_val, str) and _rv_val.startswith('='):
                    _rev_formula_template = _rv_val
                    break

        # Find the MDR column so we can default new rows to 0.01.
        _mdr_col = None
        for _mdr_name in ('MDR', 'MDR RATE', 'MDR%', 'MDR %', 'RATE'):
            if _mdr_name in header_map:
                _mdr_col = header_map[_mdr_name]
                break

        # ── Detect bold columns and right-border columns from a reference row ──
        # Scan early data rows to find which columns should be bold and which
        # carry a right-side border (e.g. the TOTAL meta-column separator).
        # Store complete Font objects (not bare Font(bold=True)) to avoid
        # producing an incomplete <font> element in styles.xml that Excel
        # flags as needing repair.
        bold_cols = set()
        bold_col_fonts = {}
        right_border_cols = {}
        for scan_r in range(2, min(total_row, 20)):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(scan_r, c)
                if cell.font and cell.font.bold:
                    bold_cols.add(c)
                    if c not in bold_col_fonts:
                        rf = cell.font
                        bold_col_fonts[c] = Font(
                            name=rf.name, size=rf.size, bold=True,
                            italic=rf.italic, strike=rf.strike,
                            color=rf.color, family=rf.family,
                            charset=rf.charset, scheme=rf.scheme,
                            underline=rf.underline, vertAlign=rf.vertAlign,
                        )
                if cell.border and cell.border.right and cell.border.right.style:
                    right_border_cols[c] = cell.border.right.style
            if bold_cols or right_border_cols:
                break

        # Map TID string -> row number for existing data rows
        tid_row_map = {}
        for r in range(2, total_row):
            val = ws.cell(r, tid_col).value
            if val is not None:
                tid_str = str(val).strip()
                if tid_str.upper() not in ('', 'NAN', 'TOTAL'):
                    tid_row_map[tid_str] = r

        modified_rows = set()
        # Track which TIDs were touched per B02 date (for zero-fill pass)
        touched_per_date = {d: set() for d in matching_b02_dates}

        # ── Apply B02 data ────────────────────────────────────────────────────
        for (tid, txn_date), amount in relevant.items():
            if txn_date not in date_col_map:
                continue
            target_col = date_col_map[txn_date]

            if tid in tid_row_map:
                r = tid_row_map[tid]
                existing = ws.cell(r, target_col).value
                try:
                    existing_val = (
                        float(existing)
                        if existing is not None
                        and str(existing).strip() not in ('', '-', 'nan')
                        else 0.0
                    )
                except (ValueError, TypeError):
                    existing_val = 0.0
                cell = ws.cell(r, target_col)
                cell.value = round(existing_val + amount, 2)
                cell.number_format = num_fmt
                modified_rows.add(r)
                stats['updated'] += 1
            else:
                # ── Insert new row just before the TOTAL row ─────────────────
                ws.insert_rows(total_row)
                new_row = total_row
                total_row += 1

                # Clear all borders copied from the TOTAL row, then re-apply
                # right-side borders and bold font that appear on real data rows.
                for c in range(1, ws.max_column + 1):
                    dst = ws.cell(new_row, c)
                    dst.border = (
                        Border(right=Side(style=right_border_cols[c]))
                        if c in right_border_cols
                        else _NO_BORDER
                    )
                    if c in bold_cols:
                        dst.font = bold_col_fonts.get(c, Font(bold=True))

                ws.cell(new_row, tid_col).value = tid

                for d, col in date_col_map.items():
                    cell = ws.cell(new_row, col)
                    cell.number_format = num_fmt
                    if d < txn_date:
                        cell.value = 0
                    elif d == txn_date:
                        cell.value = round(amount, 2)
                    # future dates: leave blank (no value, format still applied)

                # Set MDR rate for the new row so the REVENUE formula has a
                # valid multiplier (default 0.01 = 1 %).
                if _mdr_col:
                    ws.cell(new_row, _mdr_col).value = 0.01

                # Populate REVENUE cell for the new row.
                # Adjust the template formula by replacing every row-number
                # reference with new_row (REVENUE formulas only reference their
                # own row, so replacing all row numbers is safe).
                if revenue_col:
                    if _rev_formula_template:
                        _new_rev = re.sub(
                            r'(\$?[A-Z]+\$?)(\d+)',
                            lambda _m: _m.group(1) + str(new_row),
                            _rev_formula_template,
                        )
                        ws.cell(new_row, revenue_col).value = _new_rev
                    else:
                        ws.cell(new_row, revenue_col).value = 0

                tid_row_map[tid] = new_row
                modified_rows.add(new_row)
                stats['added'] += 1

            touched_per_date[txn_date].add(tid)

        # ── Zero-fill existing TIDs that did NOT transact on each B02 date ───
        # The B02 only lists terminals that transacted; all others get 0 ('-').
        for txn_date in matching_b02_dates:
            target_col = date_col_map[txn_date]
            for tid_str, r in tid_row_map.items():
                if tid_str in touched_per_date[txn_date]:
                    continue  # already set above
                cell = ws.cell(r, target_col)
                current = cell.value
                # Only fill blank / dash / NaN cells — don't overwrite real values
                if current is None or str(current).strip() in ('', '-', 'nan', '0'):
                    cell.value = 0
                    cell.number_format = num_fmt

        # ── Recalculate TOTAL meta-column for rows we modified ────────────────
        if total_meta_col:
            for r in modified_rows:
                row_total = 0.0
                for col in date_col_map.values():
                    v = ws.cell(r, col).value
                    if v is not None:
                        try:
                            row_total += float(v)
                        except (ValueError, TypeError):
                            pass
                cell = ws.cell(r, total_meta_col)
                cell.value = round(row_total, 2)
                cell.number_format = num_fmt

        # ── Recalculate TOTAL row (sum of each date column, all data rows) ──────
        for d, col in date_col_map.items():
            col_sum = 0.0
            for r in range(2, total_row):
                v = ws.cell(r, col).value
                if v is not None and not (isinstance(v, str) and v.startswith('=')):
                    try:
                        col_sum += float(v)
                    except (ValueError, TypeError):
                        pass
            ws.cell(total_row, col).value = round(col_sum, 2)

        # ── Recalculate TOTAL row's TOTAL meta-column ──────────────────────────
        # The TOTAL row's date columns were just recomputed above to reflect ALL
        # data rows (including any newly inserted ones).  Sum them directly — this
        # equals the sum of every row's TOTAL column and gives the true MTD figure.
        if total_meta_col:
            total_row_sum = 0.0
            for col in date_col_map.values():
                v = ws.cell(total_row, col).value
                if v is not None and not (isinstance(v, str) and v.startswith('=')):
                    try:
                        total_row_sum += float(v)
                    except (ValueError, TypeError):
                        pass
            ws.cell(total_row, total_meta_col).value = round(total_row_sum, 2)

        # ── REVENUE column in TOTAL row: leave original formula intact ─────────
        # Per-row REVENUE cells are =MDR*TOTAL formulas whose current values are
        # formula strings (data_only=False load).  Attempting float() on them
        # gives 0.  We skip recomputation and let Excel re-evaluate the
        # existing =SUM(H2:H...) formula on open.

        # ── Enforce continuous border on the entire TOTAL row ────────────────
        # Restore the original borders first, then guarantee a uniform top+bottom
        # border across every cell so the separator line is unbroken.
        for _rc, _rb in _saved_total_borders.items():
            ws.cell(total_row, _rc).border = _rb

        # Determine the reference top/bottom border style from the TOTAL column
        # cell (total_meta_col), which already carries the correct preset (e.g.
        # "Top and Thick Bottom Border").  Fall back to the REVENUE column, then
        # to any other cell in the row that has a border, then to thin/thick.
        def _make_side(s):
            """Rebuild a Side from an existing Side object, preserving color."""
            if s is None or not getattr(s, 'style', None):
                return None
            try:
                return Side(style=s.style, color=s.color)
            except Exception:
                return Side(style=s.style)

        _ref_b = None
        for _ref_col in filter(None, [total_meta_col, revenue_col]):
            _rb = _saved_total_borders.get(_ref_col)
            if _rb and (
                (_rb.top and _rb.top.style) or (_rb.bottom and _rb.bottom.style)
            ):
                _ref_b = _rb
                break
        if _ref_b is None:
            for _rb in _saved_total_borders.values():
                if _rb and (
                    (_rb.top and _rb.top.style) or (_rb.bottom and _rb.bottom.style)
                ):
                    _ref_b = _rb
                    break

        _top_side = _make_side(_ref_b.top)    if _ref_b else Side(style='thin')
        _bot_side = _make_side(_ref_b.bottom) if _ref_b else Side(style='thick')
        # Ensure we always have valid sides even if one half of the preset is absent
        if _top_side is None:
            _top_side = Side(style='thin')
        if _bot_side is None:
            _bot_side = Side(style='thick')

        # Apply the reference top+bottom border to every cell in the TOTAL row,
        # preserving any existing left/right cell borders.
        for _rc in range(1, ws.max_column + 1):
            _cb = ws.cell(total_row, _rc).border
            ws.cell(total_row, _rc).border = Border(
                left=_cb.left   if _cb else None,
                right=_cb.right if _cb else None,
                top=_top_side,
                bottom=_bot_side,
            )

        # ── Accumulate per-date totals for VISUALS sheet ──────────────────────
        # Read the now-recalculated TOTAL row for every date column on this sheet
        # and add to the workbook-level running totals.
        for _d, _col in date_col_map.items():
            _tv = ws.cell(total_row, _col).value
            if _tv is not None and not (isinstance(_tv, str) and _tv.startswith('=')):
                try:
                    _all_date_totals[_d] = _all_date_totals.get(_d, 0.0) + float(_tv)
                except (ValueError, TypeError):
                    pass

    # ── Update VISUALS sheet ──────────────────────────────────────────────────
    # Use case-insensitive lookup for the VISUALS sheet name.
    _visuals_sn = next(
        (s for s in wb.sheetnames if s.strip().upper() == 'VISUALS'), None
    )
    if _all_date_totals and _visuals_sn:
        _vws = wb[_visuals_sn]
        _date_row_num = None
        _dep_row_num = None
        # Scan every row; check the first few columns for the row labels because
        # some files place them in column B rather than column A.
        for _r in range(1, _vws.max_row + 1):
            for _lbl_c in range(1, min(4, _vws.max_column + 1)):
                _lbl = _vws.cell(_r, _lbl_c).value
                if _lbl is not None:
                    _lbl_u = str(_lbl).strip().upper()
                    if _lbl_u == 'DATE' and _date_row_num is None:
                        _date_row_num = _r
                    elif _lbl_u == 'DEPOSITS' and _dep_row_num is None:
                        _dep_row_num = _r
            if _date_row_num and _dep_row_num:
                break

        if _date_row_num and _dep_row_num:
            for _c in range(1, _vws.max_column + 1):
                _cv = _vws.cell(_date_row_num, _c).value
                if _cv is None:
                    continue
                _parsed_date = None
                # 1. Python datetime / date objects (openpyxl native)
                if isinstance(_cv, _dt.datetime):
                    _parsed_date = _cv.date()
                elif isinstance(_cv, _dt.date):
                    _parsed_date = _cv
                # 2. Excel serial date integer (e.g. 45123)
                elif isinstance(_cv, (int, float)) and 30000 < _cv < 70000:
                    try:
                        _parsed_date = (
                            pd.Timestamp('1899-12-30') + pd.Timedelta(days=int(_cv))
                        ).date()
                    except Exception:
                        pass
                else:
                    # 3. String — try our column-header parser first
                    try:
                        _pp = _parse_col_date(str(_cv))
                        if pd.notna(_pp):
                            _parsed_date = _pp.date()
                    except Exception:
                        pass
                    # 4. General pandas parser as last resort
                    if _parsed_date is None:
                        try:
                            _pp2 = pd.to_datetime(str(_cv), dayfirst=True, errors='coerce')
                            if pd.notna(_pp2):
                                _parsed_date = _pp2.date()
                        except Exception:
                            pass
                if _parsed_date is not None and _parsed_date in _all_date_totals:
                    _vws.cell(_dep_row_num, _c).value = round(_all_date_totals[_parsed_date], 2)

    # ── Pre-save cleanup ──────────────────────────────────────────────────────
    from openpyxl.formatting.formatting import ConditionalFormattingList as _CFList

    # Clear workbook-level defined names (print areas, print titles, named ranges)
    # that may hold stale cell-range references after insert_rows().
    try:
        for _dn in list(wb.defined_names.definedName):
            wb.defined_names.definedName.remove(_dn)
    except Exception:
        try:
            for _dn in list(wb.defined_names):
                del wb.defined_names[_dn]
        except Exception:
            pass

    for _sn in wb.sheetnames:
        _ws = wb[_sn]

        # Remove all conditional formatting rules
        _ws.conditional_formatting = _CFList()

        # Remove data validations (stale cell references after insert_rows)
        try:
            _ws.data_validations.dataValidation = []
        except Exception:
            pass

        # Remove print area / print titles (sheet-level defined names)
        try:
            _ws.print_area = None
        except Exception:
            pass
        try:
            _ws.print_title_rows = None
            _ws.print_title_cols = None
        except Exception:
            pass

        # Identify TOTAL and REVENUE column indices from the header row
        _protected_cols = set()
        for _c in range(1, _ws.max_column + 1):
            _hv = _ws.cell(1, _c).value
            if _hv is not None and str(_hv).strip().upper() in ('TOTAL', 'REVENUE'):
                _protected_cols.add(_c)

        # Clear formulas in all non-protected columns
        for _row in _ws.iter_rows(min_row=1):
            for _cell in _row:
                if _cell.column not in _protected_cols:
                    if isinstance(_cell.value, str) and _cell.value.startswith('='):
                        _cell.value = None

        # Clear the stale SheetView-level topLeftCell that pins the scroll
        # position to the last-inserted row (causing the bounce-back), then
        # freeze row 1 so the header stays visible while scrolling.
        try:
            for _sv in _ws.views.sheetView:
                _sv.topLeftCell = None
        except Exception:
            pass
        _ws.freeze_panes = 'A2'
        try:
            from openpyxl.worksheet.views import Selection as _Sel
            _sv = _ws.sheet_view
            if _sv.pane:
                _sv.pane.topLeftCell = 'A2'
            # Mutate the list in-place — direct reassignment (_sv.selection = [...])
            # can replace the Sequence descriptor's tracked list with a plain Python
            # list, which openpyxl then serialises to malformed XML and Excel rejects
            # the file as corrupted.
            _sel = _sv.selection
            if _sel is not None:
                del _sel[:]
                _sel.append(_Sel(pane='topLeft',    activeCell='A1', sqref='A1'))
                _sel.append(_Sel(pane='bottomLeft', activeCell='A2', sqref='A2'))
        except Exception:
            pass

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), stats
